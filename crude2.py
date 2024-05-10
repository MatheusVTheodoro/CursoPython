import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

class Contato:
    def __init__(self, nome, telefone):
        self.nome = nome
        self.telefone = telefone

# Função para adicionar um novo contato
def adicionar_contato():
    nome = entry_nome.get()
    telefone = entry_telefone.get()
    if nome and telefone:
        novo_contato = Contato(nome, telefone)
        contatos.append(novo_contato)
        salvar_contatos()
        atualizar_lista_contatos()
        limpar_campos()
    else:
        messagebox.showwarning("Erro", "Por favor, preencha todos os campos.")

# Função para atualizar a lista de contatos na interface
def atualizar_lista_contatos():
    lista_contatos.delete(0, tk.END)
    for contato in contatos:
        lista_contatos.insert(tk.END, f"{contato.nome}: {contato.telefone}")

# Função para limpar os campos de entrada após adicionar um novo contato
def limpar_campos():
    entry_nome.delete(0, tk.END)
    entry_telefone.delete(0, tk.END)

# Função para salvar os contatos na planilha Excel
def salvar_contatos():
    planilha = load_workbook('contatos.xlsx')
    planilha.remove(planilha.active)
    aba = planilha.create_sheet("Contatos")

    for i, contato in enumerate(contatos, start=1):
        aba[f'A{i}'] = contato.nome
        aba[f'B{i}'] = contato.telefone

    planilha.save('contatos.xlsx')

# Inicialização da janela principal
root = tk.Tk()
root.title("Lista de Contatos")

# Lista de contatos
contatos = []

# Tenta carregar os contatos da planilha, caso não exista cria uma nova
try:
    planilha = load_workbook('contatos.xlsx')
    aba = planilha['Contatos']
    for row in aba.iter_rows(values_only=True):
        contatos.append(Contato(row[0], row[1]))
except FileNotFoundError:
    Workbook().save('contatos.xlsx')

# Frame para os campos de entrada e botão de adicionar
frame_formulario = tk.Frame(root)
frame_formulario.pack(pady=10)

label_nome = tk.Label(frame_formulario, text="Nome:")
label_nome.grid(row=0, column=0, padx=5)
entry_nome = tk.Entry(frame_formulario)
entry_nome.grid(row=0, column=1, padx=5)

label_telefone = tk.Label(frame_formulario, text="Telefone:")
label_telefone.grid(row=0, column=2, padx=5)
entry_telefone = tk.Entry(frame_formulario)
entry_telefone.grid(row=0, column=3, padx=5)

btn_adicionar = tk.Button(frame_formulario, text="Adicionar", command=adicionar_contato)
btn_adicionar.grid(row=0, column=4, padx=5)

# Frame para a lista de contatos
frame_lista = tk.Frame(root)
frame_lista.pack(pady=10)

label_lista = tk.Label(frame_lista, text="Lista de Contatos:")
label_lista.pack()

scrollbar = tk.Scrollbar(frame_lista)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

lista_contatos = tk.Listbox(frame_lista, width=50, yscrollcommand=scrollbar.set)
lista_contatos.pack()

scrollbar.config(command=lista_contatos.yview)

# Atualiza a lista de contatos na inicialização
atualizar_lista_contatos()

# Execução da aplicação
root.mainloop()
