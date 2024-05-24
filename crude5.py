import tkinter as tk
from tkinter import messagebox, scrolledtext
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime

class Ticket:
    def __init__(self, cliente, problema, prioridade, data_hora):
        self.cliente = cliente
        self.problema = problema
        self.prioridade = prioridade
        self.data_hora = data_hora

# Função para adicionar um novo ticket de helpdesk
def adicionar_ticket():
    cliente = entry_cliente.get()
    problema = entry_problema.get("1.0", tk.END)  # Obtemos todo o texto do campo de texto
    prioridade = combo_prioridade.get()
    data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Data e hora atual

    if cliente and problema.strip() and prioridade:
        novo_ticket = Ticket(cliente, problema, prioridade, data_hora)
        tickets.append(novo_ticket)
        salvar_tickets()
        atualizar_lista_tickets()
        limpar_campos()
    else:
        messagebox.showwarning("Erro", "Por favor, preencha todos os campos.")

# Função para atualizar a lista de tickets na interface
def atualizar_lista_tickets():
    lista_tickets.delete(0, tk.END)
    for i, ticket in enumerate(tickets):
        lista_tickets.insert(tk.END, f"Ticket {i+1}: Cliente: {ticket.cliente} | Prioridade: {ticket.prioridade} | Data/Hora: {ticket.data_hora}")

# Função para limpar os campos de entrada após adicionar um novo ticket
def limpar_campos():
    entry_cliente.delete(0, tk.END)
    entry_problema.delete("1.0", tk.END)
    combo_prioridade.set("")

# Função para salvar os tickets na planilha Excel
def salvar_tickets():
    planilha = load_workbook('tickets.xlsx')
    planilha.remove(planilha.active)
    aba = planilha.create_sheet("Tickets")

    for i, ticket in enumerate(tickets, start=1):
        aba[f'A{i}'] = ticket.cliente
        aba[f'B{i}'] = ticket.problema
        aba[f'C{i}'] = ticket.prioridade
        aba[f'D{i}'] = ticket.data_hora

    planilha.save('tickets.xlsx')

# Função para exibir a descrição completa de um ticket
def exibir_descricao_ticket(event):
    indice_selecionado = lista_tickets.curselection()
    if indice_selecionado:
        indice_selecionado = int(indice_selecionado[0])
        ticket_selecionado = tickets[indice_selecionado]

        janela_descricao = tk.Toplevel(root)
        janela_descricao.title(f"Ticket {indice_selecionado + 1}")

        label_cliente = ttk.Label(janela_descricao, text=f"Cliente: {ticket_selecionado.cliente}")
        label_cliente.pack(pady=5)

        label_problema = ttk.Label(janela_descricao, text="Descrição do Problema:")
        label_problema.pack(pady=5)

        texto_problema = scrolledtext.ScrolledText(janela_descricao, width=40, height=10)
        texto_problema.insert(tk.END, ticket_selecionado.problema)
        texto_problema.pack(pady=5)

        label_prioridade = ttk.Label(janela_descricao, text=f"Prioridade: {ticket_selecionado.prioridade}")
        label_prioridade.pack(pady=5)

        label_data_hora = ttk.Label(janela_descricao, text=f"Data/Hora: {ticket_selecionado.data_hora}")
        label_data_hora.pack(pady=5)

# Inicialização da janela principal
root = tk.Tk()
root.title("Sistema de Helpdesk")

# Estilo
style = ttk.Style()
style.configure("TFrame", padding=10)
style.configure("TLabel", padding=5)
style.configure("TButton", padding=5)

# Lista de tickets
tickets = []

# Tenta carregar os tickets da planilha, caso não exista cria uma nova
try:
    planilha = load_workbook('tickets.xlsx')
    aba = planilha['Tickets']
    for row in aba.iter_rows(values_only=True):
        if len(row) >= 4:  # Verifica se a linha possui pelo menos 4 elementos
            tickets.append(Ticket(row[0], row[1], row[2], row[3]))
except FileNotFoundError:
    Workbook().save('tickets.xlsx')

# Frame para os campos de entrada e botão de adicionar
frame_formulario = ttk.Frame(root)
frame_formulario.pack(pady=10)

label_cliente = ttk.Label(frame_formulario, text="Cliente:")
label_cliente.grid(row=0, column=0, padx=5)
entry_cliente = ttk.Entry(frame_formulario)
entry_cliente.grid(row=0, column=1, padx=5)

label_problema = ttk.Label(frame_formulario, text="Problema:")
label_problema.grid(row=1, column=0, padx=5)
entry_problema = tk.Text(frame_formulario, height=5, width=30)
entry_problema.grid(row=1, column=1, padx=5)

label_prioridade = ttk.Label(frame_formulario, text="Prioridade:")
label_prioridade.grid(row=2, column=0, padx=5)
combo_prioridade = tk.StringVar()
combo_prioridade.set("")  # Define uma opção padrão vazia
prioridades = ["Alta", "Média", "Baixa"]
combo_prioridade_menu = ttk.OptionMenu(frame_formulario, combo_prioridade, *prioridades)
combo_prioridade_menu.grid(row=2, column=1, padx=5)

btn_adicionar = ttk.Button(frame_formulario, text="Adicionar", command=adicionar_ticket)
btn_adicionar.grid(row=3, column=0, columnspan=2, pady=5)

# Frame para a lista de tickets
frame_lista = ttk.Frame(root)
frame_lista.pack(pady=10)

label_lista = ttk.Label(frame_lista, text="Lista de Tickets:")
label_lista.pack(pady=5)

scrollbar = ttk.Scrollbar(frame_lista)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

lista_tickets = tk.Listbox(frame_lista, width=50, yscrollcommand=scrollbar.set)
lista_tickets.pack()

scrollbar.config(command=lista_tickets.yview)

# Atualiza a lista de tickets na inicialização
atualizar_lista_tickets()

# Associa a função exibir_descricao_ticket ao evento de clicar em um item da lista
lista_tickets.bind("<<ListboxSelect>>", exibir_descricao_ticket)

# Execução da aplicação
root.mainloop()
